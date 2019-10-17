import os
from rutermextract import TermExtractor
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
i=2
term_extractor = TermExtractor()
filename = input("Введите путь к файлу: ")
if not os.path.exists(filename):
    print("Указанный файл не существует")
else:
     with open(filename, encoding="utf8") as file:
         text = file.read()
     for term in term_extractor(text):
             print(term.normalized.capitalize(), term.count)        
# создаём новый файл .xlsx
     wb=Workbook()
# добавляем новый лист
     wb.create_sheet(title='Первый', index=0)
# получаем лист, с которым будем работать     
     sheet=wb['Первый']
     sheet.column_dimensions['A'].width=45
# создаём именованный стиль
     ns=NamedStyle(name='bold1')
     ns.font=Font(bold=True)
# регистрация вновь созданного стиля для дальнейшего использования
     wb.add_named_style(ns)
     sheet['A1'].style='bold1'
     sheet['B1'].style='bold1'
     sheet['A1']='Ключевое слово'
     sheet['B1']='Количество'
     for term in term_extractor(text):
         sheet.cell(row=i,column=1).value=term.normalized.capitalize()
         sheet.cell(row=i,column=2).value=term.count
         i=i+1
     wb.save('Unique.xlsx')
     print('Запись .xlsx-файла завершена.')
 
